"""Tests for the DataProfileTool: read-only structural discovery of data files/dirs.

Covers the three target scenarios:
- single file single table (CSV)
- single file multi-sheet (Excel workbook)
- multi-file discovery (directory listing of tabular files)
"""

import os

import pytest

from data_analysis_agent.tools import data_profile as dp
from data_analysis_agent.tools.data_profile import DataProfileTool


def test_schema_and_security_flags():
    tool = DataProfileTool()
    assert tool.name == "data_profile"
    # read-only discovery: safe, non-destructive, parallelizable
    assert tool.is_read_only({}) is True
    assert tool.is_destructive({}) is False
    assert tool.is_concurrency_safe({}) is True


def test_validate_requires_path():
    tool = DataProfileTool()
    assert tool.validate_input({}).valid is False
    assert tool.validate_input({"path": ""}).valid is False
    assert tool.validate_input({"path": "data.csv"}).valid is True


async def test_profiles_csv_columns_and_rows(tmp_path):
    csv = tmp_path / "sales.csv"
    csv.write_text("order_id,region,amount\n1,East,10\n2,West,20\n3,East,30\n", encoding="utf-8")
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(csv)})

    assert not result.is_error
    assert "order_id" in result.content
    assert "region" in result.content
    assert "amount" in result.content
    assert "3 rows" in result.content
    # absolute path so the model can copy it into read_csv()
    assert str(csv.resolve()) in result.content


async def test_profiles_excel_lists_all_sheets(tmp_path):
    pd = pytest.importorskip("pandas")
    pytest.importorskip("openpyxl")
    xlsx = tmp_path / "book.xlsx"
    with pd.ExcelWriter(xlsx) as writer:
        pd.DataFrame({"order_id": [1, 2], "amount": [10, 20]}).to_excel(
            writer, sheet_name="orders", index=False
        )
        pd.DataFrame({"cust_id": [1], "name": ["A"]}).to_excel(
            writer, sheet_name="customers", index=False
        )
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(xlsx)})

    assert not result.is_error
    # every sheet name surfaced so multi-sheet is discoverable, not blind-probed
    assert "orders" in result.content
    assert "customers" in result.content
    # per-sheet columns surfaced
    assert "order_id" in result.content
    assert "cust_id" in result.content


async def test_excel_metadata_has_one_table_per_sheet(tmp_path):
    pd = pytest.importorskip("pandas")
    pytest.importorskip("openpyxl")
    xlsx = tmp_path / "book.xlsx"
    with pd.ExcelWriter(xlsx) as writer:
        pd.DataFrame({"a": [1, 2], "b": [3, 4]}).to_excel(writer, sheet_name="s1", index=False)
        pd.DataFrame({"c": [5]}).to_excel(writer, sheet_name="s2", index=False)
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(xlsx)})

    prof = result.metadata["profile"]
    assert prof["kind"] == "file"
    assert prof["format"] == "excel"
    sheets = {t["sheet"]: t for t in prof["tables"]}
    assert set(sheets) == {"s1", "s2"}
    assert [c["name"] for c in sheets["s1"]["columns"]] == ["a", "b"]
    assert sheets["s1"]["n_rows_sampled"] == 2
    assert sheets["s2"]["n_rows_sampled"] == 1


async def test_profiles_directory_lists_tabular_files(tmp_path):
    pd = pytest.importorskip("pandas")
    pytest.importorskip("openpyxl")
    (tmp_path / "a.csv").write_text("x,y\n1,2\n", encoding="utf-8")
    (tmp_path / "notes.txt").write_text("hello", encoding="utf-8")
    pd.DataFrame({"k": [1]}).to_excel(tmp_path / "b.xlsx", index=False)
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(tmp_path)})

    assert not result.is_error
    assert "a.csv" in result.content
    assert "b.xlsx" in result.content
    # non-tabular files are not listed as analysable datasets
    assert "notes.txt" not in result.content


async def test_directory_metadata_lists_files(tmp_path):
    (tmp_path / "a.csv").write_text("x,y\n1,2\n", encoding="utf-8")
    (tmp_path / "c.tsv").write_text("p\tq\n1\t2\n", encoding="utf-8")
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(tmp_path)})

    prof = result.metadata["profile"]
    assert prof["kind"] == "directory"
    names = {f["name"] for f in prof["files"]}
    assert names == {"a.csv", "c.tsv"}


async def test_csv_metadata_has_structured_profile(tmp_path):
    csv = tmp_path / "s.csv"
    csv.write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(csv)})

    prof = result.metadata["profile"]
    assert prof["kind"] == "file"
    assert prof["format"] == "csv"
    assert prof["path"] == str(csv.resolve())
    table = prof["tables"][0]
    assert table["sheet"] is None
    assert [c["name"] for c in table["columns"]] == ["a", "b"]
    assert table["n_rows_sampled"] == 2


async def test_missing_path_is_error(tmp_path):
    tool = DataProfileTool(allowed_paths=[tmp_path])
    result = await tool.call({"path": str(tmp_path / "nope.csv")})
    assert result.is_error


async def test_path_outside_allowed_is_error(tmp_path):
    allowed = tmp_path / "allowed"
    allowed.mkdir()
    outside = tmp_path / "secret.csv"
    outside.write_text("a,b\n1,2\n", encoding="utf-8")
    tool = DataProfileTool(allowed_paths=[allowed])

    result = await tool.call({"path": str(outside)})

    assert result.is_error
    assert "outside" in result.content.lower()


async def test_unsupported_file_type_is_error(tmp_path):
    img = tmp_path / "pic.png"
    img.write_bytes(b"\x89PNG\r\n")
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(img)})

    assert result.is_error


# --- review-driven hardening ------------------------------------------------


async def test_directory_does_not_follow_symlink_outside_allowed(tmp_path):
    """A symlink inside an allowed dir pointing OUTSIDE must not leak (C1)."""
    allowed = tmp_path / "allowed"
    allowed.mkdir()
    secret = tmp_path / "secret.csv"
    secret.write_text("topsecret,pw\n1,2\n", encoding="utf-8")
    link = allowed / "innocent.csv"
    try:
        os.symlink(secret, link)
    except (OSError, NotImplementedError):
        pytest.skip("symlinks not supported on this platform")
    tool = DataProfileTool(allowed_paths=[allowed])

    result = await tool.call({"path": str(allowed)})

    assert not result.is_error
    assert "topsecret" not in result.content
    assert "innocent.csv" not in result.content
    names = {f["name"] for f in result.metadata["profile"]["files"]}
    assert "innocent.csv" not in names


async def test_sampled_flag_false_at_exact_threshold(tmp_path, monkeypatch):
    monkeypatch.setattr(dp, "_SAMPLE_ROWS", 3)
    csv = tmp_path / "exact.csv"
    csv.write_text("a\n1\n2\n3\n", encoding="utf-8")  # exactly 3 data rows
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(csv)})

    table = result.metadata["profile"]["tables"][0]
    assert table["n_rows_sampled"] == 3
    assert table["sampled"] is False


async def test_sampled_flag_true_above_threshold(tmp_path, monkeypatch):
    monkeypatch.setattr(dp, "_SAMPLE_ROWS", 3)
    csv = tmp_path / "big.csv"
    csv.write_text("a\n1\n2\n3\n4\n5\n", encoding="utf-8")  # 5 data rows
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(csv)})

    table = result.metadata["profile"]["tables"][0]
    assert table["sampled"] is True
    assert table["n_rows_sampled"] == 3  # reports the sample size, not the total


def test_stdlib_reader_sampled_semantics(tmp_path, monkeypatch):
    monkeypatch.setattr(dp, "_SAMPLE_ROWS", 3)
    csv = tmp_path / "x.csv"
    csv.write_text("a\n1\n2\n3\n", encoding="utf-8")
    cols, n_rows, sampled = dp._read_delimited_stdlib(csv, ",")
    assert cols == ["a"]
    assert n_rows == 3
    assert sampled is False

    csv.write_text("a\n1\n2\n3\n4\n", encoding="utf-8")
    _cols, n_rows, sampled = dp._read_delimited_stdlib(csv, ",")
    assert n_rows == 3
    assert sampled is True


async def test_empty_csv_returns_empty_profile_not_error(tmp_path):
    """pandas raises EmptyDataError on an empty file; must match stdlib (M2)."""
    csv = tmp_path / "empty.csv"
    csv.write_text("", encoding="utf-8")
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(csv)})

    assert not result.is_error
    table = result.metadata["profile"]["tables"][0]
    assert table["columns"] == []
    assert table["n_rows_sampled"] == 0


async def test_truncated_directory_header_shows_true_total(tmp_path, monkeypatch):
    monkeypatch.setattr(dp, "_MAX_DIR_FILES", 2)
    for i in range(3):
        (tmp_path / f"f{i}.csv").write_text("a\n1\n", encoding="utf-8")
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(tmp_path)})

    prof = result.metadata["profile"]
    assert prof["truncated"] is True
    assert prof["n_total"] == 3
    assert len(prof["files"]) == 2
    assert "3" in result.content
    assert "truncated" in result.content.lower()


# --- Excel header-health (P1-4.7) ------------------------------------------


def _write_messy_sheet(xlsx_path, *, title="2024 销售月报"):
    """A sheet with a title row + blank row + real header + data (header NOT at row 0)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "messy"
    ws.append([title, None, None])  # row 0: report title
    ws.append([None, None, None])  # row 1: blank
    ws.append(["date", "amount", "account"])  # row 2: real header
    ws.append(["2024-01-01", 10, "A"])
    ws.append(["2024-01-02", 20, "B"])
    wb.save(xlsx_path)


async def test_excel_detects_shifted_header(tmp_path):
    pytest.importorskip("pandas")
    pytest.importorskip("openpyxl")
    xlsx = tmp_path / "messy.xlsx"
    _write_messy_sheet(xlsx)
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(xlsx)})

    assert not result.is_error
    table = result.metadata["profile"]["tables"][0]
    assert table["header_offset"] == 2
    # profiled columns are the REAL header (not the title / Unnamed)
    col_names = [c["name"] for c in table["columns"]]
    assert col_names == ["date", "amount", "account"]
    # the warning tells the model how to re-read
    assert "header=2" in result.content
    assert "real header" in result.content


async def test_excel_clean_sheet_header_offset_zero(tmp_path):
    pd = pytest.importorskip("pandas")
    pytest.importorskip("openpyxl")
    xlsx = tmp_path / "clean.xlsx"
    with pd.ExcelWriter(xlsx) as writer:
        pd.DataFrame({"date": ["2024-01-01"], "amount": [10]}).to_excel(
            writer, sheet_name="s1", index=False
        )
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(xlsx)})

    table = result.metadata["profile"]["tables"][0]
    assert table["header_offset"] == 0
    assert "real header" not in result.content  # no false alarm on a clean sheet


async def test_excel_mixed_clean_and_messy_sheets(tmp_path):
    pd = pytest.importorskip("pandas")
    pytest.importorskip("openpyxl")
    xlsx = tmp_path / "mixed.xlsx"
    # sheet 1: clean
    with pd.ExcelWriter(xlsx) as writer:
        pd.DataFrame({"a": [1, 2]}).to_excel(writer, sheet_name="clean", index=False)
    # append a messy sheet via openpyxl
    import openpyxl

    wb = openpyxl.load_workbook(xlsx)
    ws = wb.create_sheet("messy")
    ws.append(["TITLE", None, None])
    ws.append([None, None, None])
    ws.append(["x", "y", "z"])
    ws.append([1, 2, 3])
    wb.save(xlsx)
    tool = DataProfileTool(allowed_paths=[tmp_path])

    result = await tool.call({"path": str(xlsx)})

    tables = {t["sheet"]: t for t in result.metadata["profile"]["tables"]}
    assert tables["clean"]["header_offset"] == 0
    assert tables["messy"]["header_offset"] == 2
    assert [c["name"] for c in tables["messy"]["columns"]] == ["x", "y", "z"]


def test_detect_header_offset_picks_densest_row_below_row0():
    pd = pytest.importorskip("pandas")
    # row0 title (1 non-null), row1 blank (0), row2 header (3) → offset 2
    grid = pd.DataFrame([[1.0, None, None], [None, None, None], ["a", "b", "c"]])
    assert dp._detect_header_offset(grid) == 2


def test_detect_header_offset_zero_when_row0_is_densest():
    pd = pytest.importorskip("pandas")
    grid = pd.DataFrame([["a", "b", "c"], [1, 2, 3]])  # clean: row0 is the header
    assert dp._detect_header_offset(grid) == 0


def test_detect_header_offset_zero_when_header_has_empty_cell():
    # clean header with a blank cell (2 non-null) + denser data row (3) must NOT
    # be flagged as shifted — this is the B1 false-positive vector.
    pd = pytest.importorskip("pandas")
    grid = pd.DataFrame([["name", None, "score"], ["alice", 30, 90]])
    assert dp._detect_header_offset(grid) == 0


def test_detect_header_offset_zero_when_data_is_ragged():
    # clean 2-col header + wider data row must NOT be flagged (margin < 2).
    pd = pytest.importorskip("pandas")
    grid = pd.DataFrame([["a", "b", None], [1, 2, 3]])
    assert dp._detect_header_offset(grid) == 0


def test_detect_header_offset_zero_when_header_has_two_empty_cells():
    # clean sparse header [name,None,None] (1 non-null) over dense numeric data
    # [alice,30,90] (3) — margin is 2 BUT the dense row is DATA (has ints), not a
    # header. The dtype guard must reject it (M1 vector).
    pd = pytest.importorskip("pandas")
    grid = pd.DataFrame([["name", None, None], ["alice", 30, 90], ["bob", 25, 85]])
    assert dp._detect_header_offset(grid) == 0


def test_detect_header_offset_zero_all_categorical_sparse_header():
    # all-categorical data over a sparse header, no blank row → no layout signal
    # → offset 0 (blank-row gate does not fire).
    pd = pytest.importorskip("pandas")
    grid = pd.DataFrame([["name", None, None], ["alice", "east", "A"], ["bob", "west", "B"]])
    assert dp._detect_header_offset(grid) == 0


def test_detect_header_offset_zero_clean_sparse_header_with_blank_spacer():
    # clean table: sparse header + a blank spacer row + dense numeric data.
    # The blank row is a real layout artifact, but the dense row is DATA (has
    # ints), not a header — the candidate all-string guard must reject it.
    pd = pytest.importorskip("pandas")
    grid = pd.DataFrame(
        [["name", None, None], [None, None, None], ["alice", 30, 90], ["bob", 25, 85]]
    )
    assert dp._detect_header_offset(grid) == 0


def test_detect_header_offset_fires_for_clear_shift():
    # sparse title (1) + blank (0) + dense header (3): margin >= 2 → offset 2
    pd = pytest.importorskip("pandas")
    grid = pd.DataFrame([["TITLE", None, None], [None, None, None], ["a", "b", "c"]])
    assert dp._detect_header_offset(grid) == 2
